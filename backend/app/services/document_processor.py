import csv
import hashlib
import json
import os
import re
import tempfile
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Tuple

import fitz
import markdown
import openpyxl
import PyPDF2
import pytesseract
from bs4 import BeautifulSoup
from docx import Document
from loguru import logger
from PIL import Image
from pydantic import BaseModel, Field


class DocumentType(str, Enum):
    PDF = "pdf"
    DOCX = "docx"
    XLSX = "xlsx"
    TXT = "txt"
    MD = "md"
    JSON = "json"
    CSV = "csv"
    JPG = "jpg"
    JPEG = "jpeg"
    PNG = "png"
    UNKNOWN = "unknown"


class DocumentChunk(BaseModel):
    chunk_id: str
    content: str
    document_id: str
    filename: str
    chunk_index: int
    total_chunks: int
    document_type: DocumentType
    metadata: Dict[str, Any] = Field(default_factory=dict)


class DocumentProcessor:
    def __init__(self, chunk_size: int = 1000, chunk_overlap: int = 200):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def detect_document_type(self, filename: str) -> DocumentType:
        ext = filename.lower().split(".")[-1] if "." in filename else ""
        type_map = {
            "pdf": DocumentType.PDF,
            "docx": DocumentType.DOCX,
            "xlsx": DocumentType.XLSX,
            "xls": DocumentType.XLSX,
            "txt": DocumentType.TXT,
            "md": DocumentType.MD,
            "json": DocumentType.JSON,
            "csv": DocumentType.CSV,
            "jpg": DocumentType.JPG,
            "jpeg": DocumentType.JPEG,
            "png": DocumentType.PNG,
        }
        return type_map.get(ext, DocumentType.UNKNOWN)

    def generate_document_id(self, filename: str, file_content: bytes) -> str:
        hash_input = f"{filename}{file_content[:1000]}".encode("utf-8")
        return hashlib.md5(hash_input).hexdigest()

    async def process_document(
        self,
        filename: str,
        file_content: bytes,
    ) -> Tuple[str, List[DocumentChunk]]:
        logger.info(f"Start processing document: {filename}")

        doc_type = self.detect_document_type(filename)
        if doc_type == DocumentType.UNKNOWN:
            raise ValueError(f"Unsupported document type: {filename}")

        document_id = self.generate_document_id(filename, file_content)
        text_content = await self.extract_text(filename, file_content, doc_type)
        chunks = self.chunk_text(text_content, document_id, filename, doc_type)

        logger.info(f"Document processed: {filename}, chunks={len(chunks)}")
        return document_id, chunks

    async def extract_text(
        self,
        filename: str,
        file_content: bytes,
        doc_type: DocumentType,
    ) -> str:
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{doc_type.value}") as tmp_file:
            tmp_file.write(file_content)
            tmp_path = tmp_file.name

        try:
            if doc_type == DocumentType.PDF:
                return self.extract_pdf_text(tmp_path)
            if doc_type == DocumentType.DOCX:
                return self.extract_docx_text(tmp_path)
            if doc_type == DocumentType.XLSX:
                return self.extract_xlsx_text(tmp_path)
            if doc_type == DocumentType.TXT:
                return self.extract_txt_text(tmp_path)
            if doc_type == DocumentType.MD:
                return self.extract_md_text(tmp_path)
            if doc_type == DocumentType.JSON:
                return self.extract_json_text(tmp_path)
            if doc_type == DocumentType.CSV:
                return self.extract_csv_text(tmp_path)
            if doc_type in [DocumentType.JPG, DocumentType.JPEG, DocumentType.PNG]:
                return self.extract_image_text(tmp_path)
            raise ValueError(f"Unsupported document type: {doc_type}")
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    def extract_pdf_text(self, file_path: str) -> str:
        text_parts: List[str] = []
        try:
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text_parts.append((page.extract_text() or "").strip())
        except Exception as exc:
            logger.error(f"Failed to extract PDF text: {exc}")
            raise

        extracted_text = "\n".join(part for part in text_parts if part).strip()
        if self._is_text_extraction_usable(extracted_text):
            return extracted_text

        logger.warning("PDF text extraction returned too little content, falling back to OCR")
        ocr_text = self.extract_pdf_text_with_ocr(file_path)
        return ocr_text or extracted_text

    def extract_pdf_text_with_ocr(self, file_path: str) -> str:
        page_texts: List[str] = []

        try:
            with fitz.open(file_path) as pdf_document:
                for page_index in range(pdf_document.page_count):
                    page = pdf_document.load_page(page_index)
                    pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                    image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
                    page_text = pytesseract.image_to_string(image, lang="chi_sim+eng").strip()
                    if page_text:
                        page_texts.append(page_text)
        except Exception as exc:
            logger.error(f"Failed to OCR PDF: {exc}")
            return ""

        return "\n".join(page_texts).strip()

    def _is_text_extraction_usable(self, text: str) -> bool:
        normalized = re.sub(r"\s+", "", text)
        if len(normalized) < 20:
            return False

        meaningful_chars = re.findall(r"[\u4e00-\u9fffA-Za-z0-9]", normalized)
        return len(meaningful_chars) >= max(10, int(len(normalized) * 0.3))

    def extract_docx_text(self, file_path: str) -> str:
        try:
            doc = Document(file_path)
            return "\n".join(paragraph.text for paragraph in doc.paragraphs)
        except Exception as exc:
            logger.error(f"Failed to extract DOCX text: {exc}")
            raise

    def extract_xlsx_text(self, file_path: str) -> str:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            lines: List[str] = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    if row_text:
                        lines.append(row_text)
            return "\n".join(lines)
        except Exception as exc:
            logger.error(f"Failed to extract XLSX text: {exc}")
            raise

    def extract_txt_text(self, file_path: str) -> str:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                return file.read()
        except Exception as exc:
            logger.error(f"Failed to extract TXT text: {exc}")
            raise

    def extract_md_text(self, file_path: str) -> str:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                md_content = file.read()
            html = markdown.markdown(md_content)
            soup = BeautifulSoup(html, "html.parser")
            return soup.get_text()
        except Exception as exc:
            logger.error(f"Failed to extract Markdown text: {exc}")
            raise

    def extract_json_text(self, file_path: str) -> str:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
            return json.dumps(data, ensure_ascii=False, indent=2)
        except Exception as exc:
            logger.error(f"Failed to extract JSON text: {exc}")
            raise

    def extract_csv_text(self, file_path: str) -> str:
        try:
            lines: List[str] = []
            with open(file_path, "r", encoding="utf-8") as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    lines.append(", ".join(row))
            return "\n".join(lines)
        except Exception as exc:
            logger.error(f"Failed to extract CSV text: {exc}")
            raise

    def extract_image_text(self, file_path: str) -> str:
        try:
            image = Image.open(file_path)
            return pytesseract.image_to_string(image, lang="chi_sim+eng")
        except Exception as exc:
            logger.error(f"Failed to extract OCR text: {exc}")
            raise

    def chunk_text(
        self,
        text: str,
        document_id: str,
        filename: str,
        doc_type: DocumentType,
    ) -> List[DocumentChunk]:
        logger.info(f"Start chunking document: {filename}, text_length={len(text)}")

        normalized_text = re.sub(r"\s+", " ", text).strip()
        if not normalized_text:
            logger.warning(f"Document text is empty after normalization: {filename}")
            return []

        segments = self._build_chunk_segments(normalized_text)
        chunks: List[DocumentChunk] = []

        for chunk_index, segment in enumerate(segments):
            chunks.append(
                DocumentChunk(
                    chunk_id=f"{document_id}_chunk_{chunk_index}",
                    content=segment,
                    document_id=document_id,
                    filename=filename,
                    chunk_index=chunk_index,
                    total_chunks=0,
                    document_type=doc_type,
                    metadata={
                        "char_count": len(segment),
                        "processed_at": datetime.now().isoformat(),
                    },
                )
            )

        for chunk in chunks:
            chunk.total_chunks = len(chunks)

        logger.info(f"Chunking complete: {filename}, chunks={len(chunks)}")
        return chunks

    def _build_chunk_segments(self, text: str) -> List[str]:
        if len(text) <= self.chunk_size:
            return [text]

        segments: List[str] = []
        start = 0
        text_length = len(text)

        while start < text_length:
            end = min(start + self.chunk_size, text_length)
            if end < text_length:
                boundary = self._find_chunk_boundary(text, start, end)
                if boundary > start:
                    end = boundary

            segment = text[start:end].strip()
            if segment:
                segments.append(segment)

            if end >= text_length:
                break

            next_start = end - self.chunk_overlap if self.chunk_overlap > 0 else end
            start = next_start if next_start > start else end

        return segments

    def _find_chunk_boundary(self, text: str, start: int, end: int) -> int:
        search_start = max(start, end - min(120, self.chunk_size // 3))
        boundary_chars = "\n。！？；;.!?,，、"

        for index in range(end - 1, search_start - 1, -1):
            if text[index] in boundary_chars:
                return index + 1

        return end


def create_document_processor() -> DocumentProcessor:
    from dotenv import load_dotenv

    load_dotenv()

    chunk_size = int(os.getenv("CHUNK_SIZE", "1000"))
    chunk_overlap = int(os.getenv("CHUNK_OVERLAP", "200"))

    return DocumentProcessor(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
